# @Date    : 21:31 08/06/2020
# @FileName: trans1.py
# @Software: PyCharm
from statsmodels.tsa.stattools import adfuller
import openpyxl
import time
import numpy as np
import pandas as pd
from datetime import datetime
import matplotlib.pylab as plt
import xlrd
def get_file(filepath):
    def timestamp_to_string(ts:int)->str:
        time_struct = time.localtime(ts)
        time_string = time.strftime("%Y-%m-%d-%H-%M-%S", time_struct)#
        return time_string
    wb=openpyxl.load_workbook(filepath)
    ws=wb.active
    for row_idx,col in enumerate(ws):
        if row_idx!=0:
            ws.cell(row_idx+1,1,timestamp_to_string(int(ws[row_idx+1][0].value)))
    wb.save("time_modify.xlsx")

def draw_trend(timeseries, size):
    f = plt.figure(facecolor='white')
    # 对size个数据进行移动平均
    rol_mean = timeseries.rolling(window=size).mean()
    # 对size个数据移动平均的方差
    rol_std = timeseries.rolling(window=size).std()
    timeseries.plot(color='blue', label='Original')
    rol_mean.plot(color='red', label='Rolling Mean')
    rol_std.plot(color='black', label='Rolling standard deviation')
    plt.legend(loc='best')
    plt.title('Rolling Mean & Standard Deviation')
    plt.show()


def draw_ts(timeseries):
    f = plt.figure(facecolor='white')
    timeseries.plot(color='blue')
    plt.show()

def teststationarity(ts):
    dftest = adfuller(ts)
    # 对上述函数求得的值进行语义描述
    dfoutput = pd.Series(dftest[0:4], index=['Test Statistic', 'p-value', '#Lags Used', 'Number of Observations Used'])
    for key, value in dftest[4].items():
        dfoutput['Critical Value (%s)' % key] = value
    print(dfoutput)
    #return dfoutput

if __name__ == '__main__':
    get_file("data2.xlsx")

    df = pd.read_excel('time_modify.xlsx', header=0, index_col=0)
    ts = df['vr']
    ts.index = pd.to_datetime(df.index,format="%Y-%m-%d-%H-%M-%S")
    ts=ts.replace([np.inf, -np.inf], np.nan).dropna(axis=0,how='any')
    draw_trend(ts, 12)

    draw_ts(ts)
    # Dickey-Fuller test:

    teststationarity(ts)
