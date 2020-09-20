# @Date    : 16:19 08/25/2020
# @Author  : ClassicalPi
# @FileName: preprocess.py
# @Software: PyCharm

from statsmodels.tsa.stattools import adfuller
import openpyxl
import time
import numpy as np
from scipy import stats
import pandas as pd
from datetime import datetime
import matplotlib.pylab as plt
import xlrd
from statsmodels.graphics.tsaplots import plot_acf, plot_pacf
from statsmodels.tsa.arima_model import ARIMA
import statsmodels.tsa.stattools as stattools
from statsmodels.tsa.arima_model import ARMA
from statsmodels.stats.diagnostic import unitroot_adf


def get_file(filepath):
    def timestamp_to_string(ts:int)->str:
        time_struct = time.localtime(ts)
        time_string = time.strftime("%Y-%m-%d %H:%M", time_struct)#
        return time_string
    wb=openpyxl.load_workbook(filepath)
    ws=wb.active
    for row_idx,col in enumerate(ws):
        if row_idx!=0:
            ws.cell(row_idx+1,1,timestamp_to_string(int(ws[row_idx+1][0].value)))
    wb.save("time_modify.xlsx")


if __name__ == '__main__':
    get_file("data2.xlsx")
    # df = pd.read_excel('time_modify.xlsx', header=0, index_col=0)
    # ts = df['vr']
    # print(ts.head())
    # print(ts.tail())
    # #all_missing = all_dummies.isnull().sum()
    # print(np.isnan(df))
    # ts.index = pd.to_datetime(df.index, format="%Y-%m-%d %H:%M")# .to_period('S')freq='M'
    # ts=ts.replace([np.inf, -np.inf], np.nan).dropna(axis=0, how='any')#數據清洗
    #
    # # helper = pd.DataFrame({'timestamp': pd.date_range(df['timestamp'].min(), df['timestamp'].max(), freq='M')})
    # helper = pd.DataFrame({'timestamp': pd.date_range(ts.index.min(), ts.index.max(), freq='m')})
    #
    # df = pd.merge(df, helper, on='timestamp', how='outer').sort_values('vr')
    # ts['timestamp'] = ts['timestamp'].interpolate(method='linear')
    data = pd.read_excel('time_modify.xlsx', header=0, index_col=0)
    df=pd.DataFrame()
    df['date']=data.index
    df['val']=data['vr'].values
    df['date'] = pd.to_datetime(df['date'], format="%Y-%m-%d %H:%M")
    helper = pd.DataFrame({'date': pd.date_range(df['date'].min(), df['date'].max())})

    d = pd.merge(df, helper, on='date', how='outer').sort_values('date')
    d['val'] = d['val'].interpolate(method='linear')
    print(d)