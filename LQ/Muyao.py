# @Date    : 21:31 08/06/2020
# @Author  : ClassicalPi
# @FileName: Muyao.py
# @Software: PyCharm

import openpyxl
import time

def get_file(filepath):
    def timestamp_to_string(ts:int)->str:
        time_struct = time.localtime(ts)
        time_string = time.strftime("%Y-%m-%d", time_struct)
        return time_string
    wb=openpyxl.load_workbook(filepath)
    ws=wb.active
    for row_idx,col in enumerate(ws):
        if row_idx!=0:
            ws.cell(row_idx+1,1,timestamp_to_string(int(ws[row_idx+1][0].value)))
    wb.save("time_modify.xlsx")

if __name__ == '__main__':
    get_file("data2.xlsx")